"""
BI Reporter Module
Upload CSV/Excel → Auto Clean → ML Prediction → Generate 5-Sheet Excel Report → Power BI Ready
"""
# importinf the files
import os, io, json, uuid, traceback, warnings, math
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score, mean_absolute_error
warnings.filterwarnings("ignore")

REPORT_FOLDER = Path("reports")
REPORT_FOLDER.mkdir(exist_ok=True)

SESSIONS = {}


def clean_val(v):
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if hasattr(v, 'item'):
        return v.item()
    return v


def detect_cols(df):
    rev_col = next((c for c in df.columns if any(k in c.lower() for k in ['revenue','sales','income','amount','price','turnover'])), None)
    cost_col = next((c for c in df.columns if any(k in c.lower() for k in ['cost','expense','spend','expenditure'])), None)
    date_col = next((c for c in df.columns if any(k in c.lower() for k in ['date','month','period','time','year'])), None)
    cat_col = next((c for c in df.columns if any(k in c.lower() for k in ['category','product','region','segment','department','type'])), None)
    qty_col = next((c for c in df.columns if any(k in c.lower() for k in ['qty','units','quantity','count','volume'])), None)
    return rev_col, cost_col, date_col, cat_col, qty_col


def clean_dataframe(df):
    logs = []
    orig = df.shape
    df = df.drop_duplicates()
    logs.append(f"Removed {orig[0] - len(df)} duplicate rows")
    for col in df.columns:
        miss = df[col].isna().sum()
        if miss == 0:
            continue
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].fillna(df[col].median())
        else:
            df[col] = df[col].fillna(df[col].mode()[0] if not df[col].mode().empty else 'Unknown')
        logs.append(f"'{col}': filled {miss} missing values")
    num_cols = df.select_dtypes(include=np.number).columns
    removed = 0
    for col in num_cols:
        q1, q3 = df[col].quantile(0.01), df[col].quantile(0.99)
        iqr = q3 - q1
        before = len(df)
        df = df[(df[col] >= q1 - 3*iqr) & (df[col] <= q3 + 3*iqr)]
        removed += before - len(df)
    logs.append(f"Removed {removed} outlier rows (IQR method)")
    logs.append(f"Clean shape: {len(df):,} rows × {len(df.columns)} columns")
    return df, logs


def run_ml_prediction(df, rev_col):
    if not rev_col or rev_col not in df.columns:
        return None
    try:
        num_cols = df.select_dtypes(include=np.number).columns.tolist()
        features = [c for c in num_cols if c != rev_col]
        if len(features) == 0:
            features = [rev_col]
            df_ml = df[[rev_col]].copy().dropna()
            df_ml['index'] = np.arange(len(df_ml))
            X = df_ml[['index']]
            y = df_ml[rev_col]
        else:
            X = df[features].fillna(0)
            y = df[rev_col].fillna(0)
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
        rf = RandomForestRegressor(n_estimators=100, random_state=42)
        rf.fit(X_train, y_train)
        preds = rf.predict(X_test)
        r2 = round(r2_score(y_test, preds), 4)
        mae = round(mean_absolute_error(y_test, preds), 2)
        last_row = X.iloc[[-1]]
        next_pred = round(float(rf.predict(last_row)[0]), 2)
        current_avg = round(float(y.mean()), 2)
        growth = round((next_pred - current_avg) / current_avg * 100, 2) if current_avg != 0 else 0
        fi = {}
        if hasattr(rf, 'feature_importances_'):
            fi = dict(zip(features, [round(float(v)*100, 1) for v in rf.feature_importances_]))
            fi = dict(sorted(fi.items(), key=lambda x: x[1], reverse=True)[:5])
        return {
            'r2': r2, 'mae': mae, 'next_prediction': next_pred,
            'current_avg': current_avg, 'growth_pct': growth,
            'feature_importance': fi, 'model': 'Random Forest'
        }
    except Exception as e:
        return {'error': str(e)}


def generate_excel_report(df, filename, ml_result, logs):
    wb = Workbook()
    DARK = "0D1117"
    ACCENT = "00E5FF"
    ACCENT2 = "7C5CFC"
    GREEN = "00E096"
    YELLOW = "FFD166"
    RED = "FF5C8A"
    WHITE = "E8EAF0"
    GRAY = "1E2736"
    LGRAY = "2D3748"

    def hdr_font(size=12, bold=True, color=WHITE):
        return Font(name='Calibri', size=size, bold=bold, color=color)
    def fill(color):
        return PatternFill("solid", fgColor=color)
    def center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def left():
        return Alignment(horizontal='left', vertical='center', wrap_text=True)

    rev_col, cost_col, date_col, cat_col, qty_col = detect_cols(df)
    num_cols = df.select_dtypes(include=np.number).columns.tolist()

    # SHEET 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False
    col_widths = [3, 20, 18, 18, 18, 18, 18, 3]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.merge_cells('B2:G3')
    ws1['B2'] = "BUSINESS INTELLIGENCE REPORT"
    ws1['B2'].font = Font(name='Calibri', size=20, bold=True, color=ACCENT)
    ws1['B2'].fill = fill(DARK)
    ws1['B2'].alignment = center()
    ws1.merge_cells('B4:G4')
    ws1['B4'] = f"Source: {filename} | Generated: {pd.Timestamp.now().strftime('%d %b %Y, %H:%M')}"
    ws1['B4'].font = Font(name='Calibri', size=10, color="64748B")
    ws1['B4'].fill = fill(DARK)
    ws1['B4'].alignment = center()
    kpi_row = 6
    kpis = [("Total Records", f"{len(df):,}", ACCENT), ("Columns", f"{len(df.columns)}", ACCENT2)]
    if rev_col:
        rev_series = pd.to_numeric(df[rev_col], errors='coerce').fillna(0)
        kpis.append(("Total Revenue", f"{rev_series.sum():,.0f}", GREEN))
    if cost_col and rev_col:
        cost_series = pd.to_numeric(df[cost_col], errors='coerce').fillna(0)
        profit = float(pd.to_numeric(df[rev_col], errors='coerce').fillna(0).sum()) - float(cost_series.sum())
        kpis.append(("Net Profit", f"{profit:,.0f}", GREEN if profit > 0 else RED))
    kpi_cols = ['B', 'C', 'D', 'E', 'F', 'G']
    for i, (label, value, color) in enumerate(kpis[:6]):
        col = kpi_cols[i]
        ws1[f'{col}{kpi_row}'] = label
        ws1[f'{col}{kpi_row}'].font = Font(name='Calibri', size=9, color="94A3B8")
        ws1[f'{col}{kpi_row}'].fill = fill(LGRAY)
        ws1[f'{col}{kpi_row}'].alignment = center()
        ws1[f'{col}{kpi_row+1}'] = value
        ws1[f'{col}{kpi_row+1}'].font = Font(name='Calibri', size=16, bold=True, color=color)
        ws1[f'{col}{kpi_row+1}'].fill = fill(LGRAY)
        ws1[f'{col}{kpi_row+1}'].alignment = center()

    # SHEET 2: Cleaned Data
    ws2 = wb.create_sheet("Cleaned Data")
    for ci, col in enumerate(df.columns, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.font = Font(name='Calibri', size=10, bold=True, color=ACCENT)
        cell.fill = fill(GRAY)
        cell.alignment = center()
        ws2.column_dimensions[get_column_letter(ci)].width = max(14, len(str(col)) + 4)
    display_df = df.head(50000)
    for ri, row in enumerate(display_df.itertuples(index=False), 2):
        bg = DARK if ri % 2 == 0 else "0A0C10"
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val if not (isinstance(val, float) and math.isnan(val)) else None)
            cell.font = Font(name='Calibri', size=9, color="94A3B8")
            cell.fill = fill(bg)

    # SHEET 3: Statistics
    ws3 = wb.create_sheet("Statistics")
    stat_headers = ["Column", "Min", "Max", "Mean", "Median", "Std Dev", "Sum", "Count"]
    for ci, h in enumerate(stat_headers, 2):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.font = Font(name='Calibri', size=10, bold=True, color=ACCENT)
        cell.fill = fill(GRAY)
        ws3.column_dimensions[get_column_letter(ci)].width = 16
    for ri, col in enumerate(num_cols, 3):
        s = df[col].dropna()
        vals = [col, round(s.min(),2), round(s.max(),2), round(s.mean(),2), round(s.median(),2), round(s.std(),2), round(s.sum(),2), len(s)]
        for ci, v in enumerate(vals, 2):
            ws3.cell(row=ri, column=ci, value=v).font = Font(name='Calibri', size=10, color=WHITE)
            ws3.cell(row=ri, column=ci).fill = fill(DARK)

    # SHEET 4: Charts
    ws4 = wb.create_sheet("Charts")
    if rev_col and cat_col and cat_col in df.columns:
        grp = df.groupby(cat_col)[rev_col].sum().reset_index().sort_values(rev_col, ascending=False).head(10)
        ws4['B1'] = cat_col
        ws4['C1'] = f"Total {rev_col}"
        for i, (_, row) in enumerate(grp.iterrows(), 2):
            ws4.cell(row=i, column=2, value=str(row[cat_col]))
            ws4.cell(row=i, column=3, value=round(float(row[rev_col]), 2))
        bar = BarChart()
        bar.title = f"Revenue by {cat_col}"
        bar.width = 22
        bar.height = 14
        data_ref = Reference(ws4, min_col=3, min_row=1, max_row=len(grp)+1)
        cats_ref = Reference(ws4, min_col=2, min_row=2, max_row=len(grp)+1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        ws4.add_chart(bar, "E2")

    # SHEET 5: Power BI
    ws5 = wb.create_sheet("Power BI Ready")
    ws5.merge_cells('B2:H2')
    ws5['B2'] = "POWER BI INTEGRATION GUIDE"
    ws5['B2'].font = Font(name='Calibri', size=14, bold=True, color=ACCENT2)
    ws5['B2'].fill = fill(DARK)
    ws5['B2'].alignment = center()
    steps = [
        ("Step 1", "Open Power BI Desktop"),
        ("Step 2", "Click 'Get Data' → 'Excel Workbook'"),
        ("Step 3", "Select this file and choose 'Cleaned Data' sheet"),
        ("Step 4", "Click 'Load' → Your data is imported!"),
        ("Step 5", "Create visuals: Bar, Line, Pie, Map — all ready!"),
    ]
    for r5, (step, desc) in enumerate(steps, 4):
        ws5[f'B{r5}'] = step
        ws5[f'C{r5}'] = desc
        ws5[f'B{r5}'].font = Font(name='Calibri', size=10, bold=True, color=ACCENT)
        ws5[f'C{r5}'].font = Font(name='Calibri', size=10, color=WHITE)
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 60

    report_path = REPORT_FOLDER / f"BI_Report_{uuid.uuid4().hex[:8]}.xlsx"
    wb.save(report_path)
    return report_path


def run_pipeline(file_bytes, filename):
    if filename.endswith('.csv'):
        df = pd.read_csv(io.BytesIO(file_bytes), low_memory=False)
    else:
        df = pd.read_excel(io.BytesIO(file_bytes))
    orig_shape = df.shape
    df, logs = clean_dataframe(df)
    rev_col, cost_col, date_col, cat_col, qty_col = detect_cols(df)
    if rev_col and cost_col and cost_col in df.columns:
        df[rev_col] = pd.to_numeric(df[rev_col], errors='coerce').fillna(0)
        df[cost_col] = pd.to_numeric(df[cost_col], errors='coerce').fillna(0)
        df['Profit'] = df[rev_col] - df[cost_col]
        df['Profit_Margin_%'] = (df['Profit'] / df[rev_col].replace(0, float('nan')) * 100).round(2).fillna(0)
        logs.append("Engineered 'Profit' and 'Profit_Margin_%' columns")
    ml_result = run_ml_prediction(df, rev_col)
    report_path = generate_excel_report(df, filename, ml_result, logs)
    num_cols = df.select_dtypes(include=np.number).columns.tolist()
    stats = {}
    for col in num_cols[:10]:
        s = df[col].dropna()
        stats[col] = {
            'min': clean_val(round(float(s.min()), 2)),
            'max': clean_val(round(float(s.max()), 2)),
            'mean': clean_val(round(float(s.mean()), 2)),
            'sum': clean_val(round(float(s.sum()), 2)),
        }
    cat_data = {}
    if cat_col and cat_col in df.columns and rev_col:
        grp = df.groupby(cat_col)[rev_col].sum().sort_values(ascending=False).head(8)
        cat_data = {'labels': list(grp.index.astype(str)), 'values': [clean_val(v) for v in grp.values]}
    preview = df.head(100).replace({float('nan'): None, float('inf'): None, float('-inf'): None}).to_dict(orient='records')
    return {
        'original_shape': list(orig_shape), 'clean_shape': [len(df), len(df.columns)],
        'columns': list(df.columns), 'logs': logs, 'stats': stats,
        'cat_data': cat_data,
        'ml': {k: clean_val(v) if not isinstance(v, dict) else {kk: clean_val(vv) for kk, vv in v.items()} for k, v in (ml_result or {}).items()},
        'preview': preview, 'report_path': str(report_path),
        'detected': {'revenue': rev_col, 'cost': cost_col, 'category': cat_col, 'date': date_col},
    }


def process_upload(file_bytes, filename):
    result = run_pipeline(file_bytes, filename)
    job_id = str(uuid.uuid4())
    SESSIONS[job_id] = result
    return job_id, result


def get_results(job_id):
    return SESSIONS.get(job_id)


def get_download_path(job_id):
    if job_id not in SESSIONS:
        return None
    return Path(SESSIONS[job_id]['report_path'])
